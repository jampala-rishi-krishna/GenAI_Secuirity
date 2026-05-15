"""
OWASP LLM Security Test Runner
================================
Install:  pip install requests openpyxl tqdm

Usage:
  python run_tests.py --file test_cases.xlsx
  python run_tests.py --file test_cases.xlsx --url http://localhost:8000
  python run_tests.py --file test_cases.xlsx --category PromptInjection
  python run_tests.py --file test_cases.xlsx --dry-run
  python run_tests.py --file test_cases.xlsx --workers 5 --delay 0.2
  python run_tests.py --file test_cases.xlsx --start-row 250   (resume)

Flags:
  --file          Path to test_cases.xlsx           (required)
  --url           API base URL                       (default: http://localhost:8000)
  --category      Run only this category             (optional)
  --dry-run       Validate Excel only, no API calls
  --workers       Parallel threads 1-10              (default: 3)
  --delay         Seconds between batches            (default: 0.3)
  --timeout       Per-request timeout secs           (default: 30)
  --start-row     Resume from this Excel row         (default: 2)
  --output        Output xlsx filename               (default: overwrites --file)
  --summary-file  Summary xlsx filename              (default: test_results_summary.xlsx)
"""

import argparse
import json
import sys
import time
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path

import requests
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
PASS_FILL = PatternFill("solid", start_color="C6EFCE")
FAIL_FILL = PatternFill("solid", start_color="FFC7CE")
FP_FILL   = PatternFill("solid", start_color="FFEB9C")
FN_FILL   = PatternFill("solid", start_color="FF9900")
ERR_FILL  = PatternFill("solid", start_color="E0E0E0")
PASS_FONT = Font(color="006100", name="Calibri", size=9, bold=True)
FAIL_FONT = Font(color="9C0006", name="Calibri", size=9, bold=True)
DEF_FONT  = Font(name="Calibri", size=9)
_T  = Side(style="thin", color="CCCCCC")
BDR = Border(left=_T, right=_T, top=_T, bottom=_T)

RES_HDRS   = ["test_id","category","input_text",
              "expected_threat_detected","expected_threat_type","expected_http_status",
              "actual_http_status","actual_threat_detected","actual_threat_type",
              "response_time_ms","sanitized_input","actual_ai_response","full_response_json",
              "PASS_FAIL","failure_reason","timestamp"]
RES_WIDTHS = [10,20,60,22,28,18,16,22,28,16,40,80,60,12,40,22]

VER_HDRS   = ["test_id","category","owasp_ref","INPUT (sent to API)","OUTPUT (AI response received)",
              "threat_detected","threat_type","security_level","PASS_FAIL","notes"]
VER_WIDTHS = [10,20,14,70,80,18,28,16,12,40]


# ── Args ──────────────────────────────────────────────────────────────────────
def parse_args():
    p = argparse.ArgumentParser(description="OWASP LLM Security Test Runner",
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--file",         required=True)
    p.add_argument("--url",          default="http://localhost:8000")
    p.add_argument("--category",     default=None)
    p.add_argument("--dry-run",      action="store_true")
    p.add_argument("--workers",      type=int,   default=3)
    p.add_argument("--delay",        type=float, default=0.3)
    p.add_argument("--timeout",      type=int,   default=30)
    p.add_argument("--start-row",    type=int,   default=2)
    p.add_argument("--output",       default="")
    p.add_argument("--summary-file", default="test_results_summary.xlsx")
    return p.parse_args()


# ── API ───────────────────────────────────────────────────────────────────────
def call_api(base_url, text, timeout):
    t0 = time.time()
    try:
        r = requests.post(
            f"{base_url}/api/v1/security/test",
            json={"test_type": "input_validation", "test_input": text},
            timeout=timeout,
        )
        ms = int((time.time() - t0) * 1000)
        try:
            body = r.json()
        except Exception:
            body = {}
        return {"status": r.status_code, "body": body, "ms": ms}
    except requests.exceptions.Timeout:
        return {"status": 504, "body": {}, "ms": timeout * 1000}
    except requests.exceptions.ConnectionError:
        return {"status": 0, "body": {}, "ms": 0}
    except Exception:
        return {"status": 0, "body": {}, "ms": 0}


def parse_threat(body):
    """Return (threat_detected, threat_type, sanitized, ai_response)."""
    ai_resp = str(body.get("ai_response") or "")
    # Updated endpoint shape (after Task 3 fix)
    if "threat_detected" in body:
        return (bool(body["threat_detected"]),
                str(body.get("threat_type") or "None"),
                str(body.get("sanitized_input") or ""),
                ai_resp)
    # Original nested shape fallback
    report    = (body.get("security_report") or {}) if isinstance(body, dict) else {}
    input_sec = (report.get("input_security") or {}) if isinstance(report, dict) else {}
    threats   = input_sec.get("threats_detected") or []
    if isinstance(threats, list) and threats:
        return True, str(threats[0]), str(input_sec.get("sanitized_input") or ""), ai_resp
    if isinstance(threats, str) and threats:
        return True, threats, "", ai_resp
    level = str(body.get("security_level") or "low").lower()
    return level in ("medium","high","critical"), str(body.get("threat_type") or "None"), "", ai_resp


def get_verdict(cat, exp_threat, exp_http, act_http, act_threat, act_type):
    if act_http == 0:
        return "ERROR", "No connection — is the backend running?"
    if act_http >= 500:
        return "ERROR", f"Server error {act_http}"
    if act_http == 429:
        return "ERROR", "Rate limited (429) — increase --delay or reduce --workers"

    exp_bool = str(exp_threat).strip().lower() == "true"
    http_ok  = act_http == exp_http or (exp_http == 200 and act_http in (200, 201))

    if cat == "AuthBypass":
        return ("PASS", "") if act_http in (401, 403) or act_threat \
               else ("FAIL", f"Auth bypass not blocked — HTTP {act_http}")

    if cat == "RateLimit":
        return ("PASS", "") if act_http in (200, 400, 429) \
               else ("FAIL", f"Unexpected HTTP {act_http}")

    if cat == "NormalChat":
        if act_threat:
            return "FAIL", "FALSE POSITIVE — benign query flagged as threat"
        return ("PASS", "") if http_ok else ("FAIL", f"HTTP {act_http} ≠ {exp_http}")

    if exp_bool:
        return ("PASS", "") if act_threat \
               else ("FAIL", "FALSE NEGATIVE — attack not detected")

    return ("PASS", "") if http_ok else ("FAIL", f"HTTP {act_http} ≠ {exp_http}")


# ── Excel helpers ─────────────────────────────────────────────────────────────
def ensure_results_sheet(wb):
    if "Results" not in wb.sheetnames:
        ws = wb.create_sheet("Results")
    else:
        ws = wb["Results"]
    if ws.max_row <= 1:
        for i, w in enumerate(RES_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c, h in enumerate(RES_HDRS, 1):
            cell = ws.cell(1, c, h)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = BDR
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(RES_HDRS))}1"
    return ws


def write_row(ws, row_idx, cols, verd, reason):
    # wrap_text columns: input_text(3), actual_ai_response(12), full_response_json(13)
    wrap_cols = {3, 12, 13}
    for c, val in enumerate(cols, 1):
        safe = str(val)[:32767] if val is not None else ""
        cell = ws.cell(row_idx, c, safe)
        cell.font   = DEF_FONT
        cell.border = BDR
        cell.alignment = Alignment(vertical="top", wrap_text=(c in wrap_cols))
    ws.row_dimensions[row_idx].height = 18

    # PASS_FAIL col is now index 14, failure_reason is 15
    vc = ws.cell(row_idx, 14)
    fc = ws.cell(row_idx, 15)
    if verd == "PASS":
        vc.fill = PASS_FILL; vc.font = PASS_FONT
    elif "FALSE POSITIVE" in reason:
        vc.fill = FP_FILL;   vc.font = FAIL_FONT; fc.fill = FP_FILL
    elif "FALSE NEGATIVE" in reason:
        vc.fill = FN_FILL;   vc.font = FAIL_FONT; fc.fill = FN_FILL
    elif verd == "FAIL":
        vc.fill = FAIL_FILL; vc.font = FAIL_FONT
    else:
        vc.fill = ERR_FILL


# ── Verification sheet ────────────────────────────────────────────────────────
ORANGE_FILL = PatternFill("solid", start_color="FCE4D6")
BLUE_FILL   = PatternFill("solid", start_color="DDEEFF")

def ensure_verification_sheet(wb):
    if "InputOutput" not in wb.sheetnames:
        ws = wb.create_sheet("InputOutput")
    else:
        ws = wb["InputOutput"]
    if ws.max_row <= 1:
        for i, w in enumerate(VER_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for c, h in enumerate(VER_HDRS, 1):
            cell = ws.cell(1, c, h)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = BDR
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(VER_HDRS))}1"
    return ws


def write_verification_row(ws, row_idx, tc, ai_resp, td, tt, sec_level, verd, reason, owasp=""):
    # alternating light background per category for readability
    bg = ORANGE_FILL if row_idx % 2 == 0 else BLUE_FILL

    vals = [
        tc["test_id"],
        tc["category"],
        owasp,
        tc["input"],          # full input
        ai_resp,              # full AI output
        str(td),              # threat_detected
        tt,                   # threat_type
        sec_level,
        verd,
        reason,
    ]
    for c, val in enumerate(vals, 1):
        safe = str(val)[:32767] if val is not None else ""
        cell = ws.cell(row_idx, c, safe)
        cell.font   = DEF_FONT
        cell.border = BDR
        # wrap input (col 4) and output (col 5)
        cell.alignment = Alignment(vertical="top", wrap_text=(c in (4, 5, 10)))
        if c not in (9,):  # don't override verdict color
            cell.fill = bg

    # color the PASS_FAIL cell (col 9)
    vc = ws.cell(row_idx, 9)
    if verd == "PASS":
        vc.fill = PASS_FILL; vc.font = PASS_FONT
    elif "FALSE" in reason:
        vc.fill = FN_FILL;   vc.font = FAIL_FONT
    elif verd == "FAIL":
        vc.fill = FAIL_FILL; vc.font = FAIL_FONT
    else:
        vc.fill = ERR_FILL

    ws.row_dimensions[row_idx].height = 60  # taller rows for readable I/O


# ── Summary file ──────────────────────────────────────────────────────────────
IO_HDRS   = ["#","test_id","category","owasp_ref",
             "INPUT sent to API","EXPECTED threat","EXPECTED type",
             "AI RESPONSE received","threat_detected","threat_type",
             "security_level","PASS / FAIL","failure_reason","response_ms"]
IO_WIDTHS = [5, 10, 18, 12, 72, 16, 22, 80, 16, 22, 14, 12, 40, 12]

def save_summary(stats, path, src_file, all_results=None):
    wb2 = Workbook()

    # ── Sheet 1: Category summary ──────────────────────────────────────────────
    ws  = wb2.active
    ws.title = "CategorySummary"
    for i, w in enumerate([28,10,10,10,10,8,8,10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    hdrs = ["Category","Total","Pass","Fail","Pass%","FP","FN","Errors"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center"); cell.border = BDR

    totals = defaultdict(int)
    row = 2
    for cat in sorted(stats):
        s = stats[cat]
        rate = round(s["pass"] / max(s["total"], 1) * 100, 1)
        for c, v in enumerate([cat, s["total"], s["pass"], s["fail"],
                                rate, s["fp"], s["fn"], s["error"]], 1):
            cell = ws.cell(row, c, v)
            cell.font = DEF_FONT; cell.border = BDR
            cell.alignment = Alignment(horizontal="left" if c==1 else "center")
            if verd_color := (PASS_FILL if (c==3 and v==s["total"]) else None):
                cell.fill = verd_color
        for k in ("total","pass","fail","error","fp","fn"):
            totals[k] += s[k]
        row += 1

    rate_t = round(totals["pass"] / max(totals["total"], 1) * 100, 1)
    for c, v in enumerate(["TOTAL", totals["total"], totals["pass"], totals["fail"],
                            rate_t, totals["fp"], totals["fn"], totals["error"]], 1):
        cell = ws.cell(row, c, v)
        cell.font = Font(name="Calibri", size=9, bold=True, color="1F4E79")
        cell.fill = PatternFill("solid", start_color="D6E4F0")
        cell.border = BDR
        cell.alignment = Alignment(horizontal="left" if c==1 else "center")

    # ── Sheet 2: InputOutput — every test with input + response ───────────────
    ws_io = wb2.create_sheet("InputOutput")
    for i, w in enumerate(IO_WIDTHS, 1):
        ws_io.column_dimensions[get_column_letter(i)].width = w
    for c, h in enumerate(IO_HDRS, 1):
        cell = ws_io.cell(1, c, h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = BDR
    ws_io.row_dimensions[1].height = 32
    ws_io.freeze_panes = "A2"
    ws_io.auto_filter.ref = f"A1:{get_column_letter(len(IO_HDRS))}1"

    if all_results:
        for idx, r in enumerate(all_results, 1):
            bg = ORANGE_FILL if idx % 2 == 0 else BLUE_FILL
            vals = [
                idx,
                r.get("test_id",""),
                r.get("category",""),
                r.get("owasp",""),
                r.get("input",""),           # full INPUT prompt
                r.get("exp_threat",""),      # expected threat
                r.get("exp_type",""),        # expected threat type
                r.get("ai_response",""),     # AI RESPONSE received
                str(r.get("threat_detected","")),
                r.get("threat_type",""),
                r.get("security_level",""),
                r.get("verdict",""),
                r.get("reason",""),
                r.get("ms", ""),
            ]
            data_row = idx + 1
            for c, val in enumerate(vals, 1):
                safe = str(val)[:32767] if val is not None else ""
                cell = ws_io.cell(data_row, c, safe)
                cell.font   = DEF_FONT
                cell.border = BDR
                cell.alignment = Alignment(vertical="top",
                                           wrap_text=(c in (5, 8, 13)))
                cell.fill = bg

            # Color verdict cell (col 12)
            vc = ws_io.cell(data_row, 12)
            verd = r.get("verdict","")
            reason = r.get("reason","")
            if verd == "PASS":
                vc.fill = PASS_FILL; vc.font = PASS_FONT
            elif "FALSE POSITIVE" in reason:
                vc.fill = FP_FILL;   vc.font = FAIL_FONT
            elif "FALSE NEGATIVE" in reason:
                vc.fill = FN_FILL;   vc.font = FAIL_FONT
            elif verd == "FAIL":
                vc.fill = FAIL_FILL; vc.font = FAIL_FONT
            else:
                vc.fill = ERR_FILL

            ws_io.row_dimensions[data_row].height = 55

    # ── Sheet 3: RunInfo ───────────────────────────────────────────────────────
    ws3 = wb2.create_sheet("RunInfo")
    ws3.column_dimensions["A"].width = 30
    ws3.column_dimensions["B"].width = 50
    for r2, (k, v) in enumerate([
        ("Run timestamp",        datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Source file",          src_file),
        ("Total tests run",      totals["total"]),
        ("Overall pass rate %",  rate_t),
        ("Passed",               totals["pass"]),
        ("Failed",               totals["fail"]),
        ("Errors",               totals["error"]),
        ("False positives (FP)", totals["fp"]),
        ("False negatives (FN)", totals["fn"]),
    ], 1):
        ws3.cell(r2, 1, k).font  = Font(name="Calibri", size=9, bold=True)
        ws3.cell(r2, 2, str(v)).font = Font(name="Calibri", size=9)

    wb2.save(path)
    return dict(totals)


# ── Dry-run ───────────────────────────────────────────────────────────────────
def run_dry(wb, args):
    ws = wb["test_cases"]
    by_cat, issues = defaultdict(int), []
    for r in range(2, ws.max_row + 1):
        cat = str(ws.cell(r, 2).value or "").strip()
        inp = str(ws.cell(r, 4).value or "").strip()
        det = str(ws.cell(r, 5).value or "").strip()
        by_cat[cat] += 1
        if not inp: issues.append(f"Row {r}: empty input_text")
        if det not in ("True","False"):
            issues.append(f"Row {r}: expected_threat_detected={det!r}")
    print(f"\n{'─'*58}")
    print(f"  DRY RUN  ·  {args.file}")
    print(f"{'─'*58}")
    print(f"  Rows:  {ws.max_row - 1}   Endpoint: POST {args.url}/api/v1/security/test")
    print(f"  Workers: {args.workers}   Delay: {args.delay}s\n  Distribution:")
    for cat, cnt in sorted(by_cat.items()):
        mark = "  ← will run" if (not args.category or cat == args.category) else ""
        print(f"    {cat:<26} {cnt:>5}{mark}")
    if issues:
        print(f"\n  ⚠  {len(issues)} issue(s):")
        for i in issues[:15]: print(f"    {i}")
    else:
        print(f"\n  ✅  Validation passed — no issues found")
    print()


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    args = parse_args()
    args.workers = max(1, min(10, args.workers))
    out_file = args.output or args.file

    print(f"\n📂  Loading {args.file} ...")
    if not Path(args.file).exists():
        print(f"❌  Not found: {args.file}"); sys.exit(1)
    wb = load_workbook(args.file)
    if "test_cases" not in wb.sheetnames:
        print("❌  Sheet 'test_cases' not found"); sys.exit(1)

    if args.dry_run:
        run_dry(wb, args); return

    # Health check
    print(f"🔗  Checking backend {args.url} ...")
    try:
        hc = requests.get(f"{args.url}/health", timeout=5)
        print(f"   {'✅' if hc.status_code==200 else '⚠ '} /health → {hc.status_code}")
    except Exception as e:
        print(f"   ❌  {e}")
        print("   Run:  uvicorn main:app --reload --host 0.0.0.0 --port 8000")
        sys.exit(1)

    ws_tc  = wb["test_cases"]
    ws_res = ensure_results_sheet(wb)
    ws_ver = ensure_verification_sheet(wb)

    # Collect rows — also grab owasp_reference (col 8) if present
    rows = []
    for r in range(args.start_row, ws_tc.max_row + 1):
        cat = str(ws_tc.cell(r, 2).value or "").strip()
        if args.category and cat != args.category:
            continue
        inp = str(ws_tc.cell(r, 4).value or "").strip()
        if not inp:
            continue
        rows.append({
            "test_id":   str(ws_tc.cell(r, 1).value or f"TC_{r}"),
            "category":  cat,
            "input":     inp,
            "exp_threat":str(ws_tc.cell(r, 5).value or "False"),
            "exp_type":  str(ws_tc.cell(r, 6).value or "None"),
            "exp_http":  int(ws_tc.cell(r, 7).value or 200),
            "owasp":     str(ws_tc.cell(r, 8).value or ""),
        })

    total = len(rows)
    fmsg  = f" (category={args.category})" if args.category else ""
    print(f"\n🧪  {total} tests{fmsg}  [{args.workers} workers · {args.delay}s delay]\n")

    # Shared state (lock-protected)
    lock        = threading.Lock()
    res_row     = [ws_res.max_row + 1]
    ver_row     = [ws_ver.max_row + 1]
    stats       = defaultdict(lambda: {"total":0,"pass":0,"fail":0,"error":0,"fp":0,"fn":0})
    all_results = []          # ← collects every row for the summary InputOutput sheet
    done        = [0]
    next_save   = [50]

    pbar = tqdm(total=total, unit="test", ncols=80) if HAS_TQDM else None

    def process(tc):
        res = call_api(args.url, tc["input"], args.timeout)
        td, tt, si, ai_resp = parse_threat(res["body"])
        sec_level = str(res["body"].get("security_level") or "low") if isinstance(res["body"], dict) else "low"
        verd, reason = get_verdict(
            tc["category"], tc["exp_threat"], tc["exp_http"],
            res["status"], td, tt,
        )
        # Results sheet row (full detail)
        cols = [
            tc["test_id"], tc["category"], tc["input"][:200],
            tc["exp_threat"], tc["exp_type"], tc["exp_http"],
            res["status"], str(td), tt,
            res["ms"], si[:200], ai_resp[:500], json.dumps(res["body"])[:300],
            verd, reason,
            datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        ]
        return tc["category"], verd, reason, cols, tc, td, tt, sec_level, ai_resp

    batch = args.workers
    for start in range(0, total, batch):
        chunk = rows[start: start + batch]
        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            fmap = {ex.submit(process, tc): tc for tc in chunk}
            for fut in as_completed(fmap):
                try:
                    cat, verd, reason, cols, tc, td, tt, sec_level, ai_resp = fut.result()
                except Exception as e:
                    tc  = fmap[fut]
                    cat = tc["category"]
                    verd, reason = "ERROR", str(e)
                    cols = [tc["test_id"], cat, tc["input"][:200],
                            tc["exp_threat"], tc["exp_type"], tc["exp_http"],
                            0,"False","None",0,"","","",verd,reason,
                            datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")]
                    td, tt, sec_level, ai_resp = False, "None", "low", ""

                with lock:
                    s = stats[cat]
                    s["total"] += 1
                    if verd == "PASS":
                        s["pass"] += 1
                    elif verd == "FAIL":
                        s["fail"] += 1
                        if "FALSE POSITIVE" in reason: s["fp"] += 1
                        if "FALSE NEGATIVE" in reason: s["fn"] += 1
                    else:
                        s["error"] += 1

                    write_row(ws_res, res_row[0], cols, verd, reason)
                    res_row[0] += 1

                    write_verification_row(
                        ws_ver, ver_row[0], tc, ai_resp,
                        td, tt, sec_level, verd, reason, tc.get("owasp","")
                    )
                    ver_row[0] += 1

                    # Save individual result for summary InputOutput sheet
                    all_results.append({
                        "test_id":        tc.get("test_id",""),
                        "category":       cat,
                        "owasp":          tc.get("owasp",""),
                        "input":          tc.get("input",""),
                        "exp_threat":     tc.get("exp_threat",""),
                        "exp_type":       tc.get("exp_type",""),
                        "ai_response":    ai_resp,
                        "threat_detected":td,
                        "threat_type":    tt,
                        "security_level": sec_level,
                        "verdict":        verd,
                        "reason":         reason,
                        "ms":             cols[9] if len(cols) > 9 else "",
                    })

                    done[0] += 1
                    if done[0] >= next_save[0]:
                        wb.save(out_file)
                        next_save[0] += 50

                if pbar:
                    p = sum(s["pass"] for s in stats.values())
                    f = sum(s["fail"] for s in stats.values())
                    pbar.update(1)
                    pbar.set_postfix(P=p, F=f)
                elif done[0] % 25 == 0:
                    p = sum(s["pass"] for s in stats.values())
                    print(f"  [{done[0]}/{total}] pass={p}")

        if args.delay > 0:
            time.sleep(args.delay)

    if pbar:
        pbar.close()

    wb.save(out_file)
    totals = save_summary(stats, args.summary_file, args.file, all_results)

    # Print summary table
    p_all = totals["pass"]; f_all = totals["fail"]
    t_all = totals["total"]; e_all = totals["error"]
    rate  = round(p_all / max(t_all, 1) * 100, 1)
    print(f"\n{'═'*72}")
    print(f"  {'CATEGORY':<24} {'TOTAL':>6} {'PASS':>6} {'FAIL':>6}"
          f" {'PASS%':>7} {'FP':>5} {'FN':>5}")
    print(f"  {'─'*66}")
    for cat in sorted(stats):
        s = stats[cat]
        r = round(s["pass"] / max(s["total"],1) * 100, 1)
        print(f"  {cat:<24} {s['total']:>6} {s['pass']:>6} {s['fail']:>6}"
              f" {r:>6.1f}% {s['fp']:>5} {s['fn']:>5}")
    print(f"  {'═'*66}")
    print(f"  {'TOTAL':<24} {t_all:>6} {p_all:>6} {f_all:>6}"
          f" {rate:>6.1f}% {totals['fp']:>5} {totals['fn']:>5}")
    print(f"  Errors: {e_all}")
    print(f"{'═'*72}")
    print(f"\n  📊  Results  → {out_file}  (Results sheet)")
    print(f"  📋  Summary  → {args.summary_file}\n")


if __name__ == "__main__":
    main()
